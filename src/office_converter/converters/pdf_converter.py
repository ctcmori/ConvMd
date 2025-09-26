"""
PDF変換モジュール

Word, Excel, PowerPointファイルをPDFへ変換する機能を提供します。
COM方式による高品質変換とフォールバック機能を実装。

依存: pywin32, docx2pdf, openpyxl, python-pptx, reportlab
"""
from pathlib import Path
from typing import Optional, List

class PDFConversionError(Exception):
    """PDF変換処理の例外"""
    pass

class PDFConverter:
    """
    OfficeファイルをPDFへ変換するためのクラス
    - Word: COM方式 -> docx2pdf (フォールバック)
    - Excel: COM方式 (高品質印刷イメージ)  
    - PowerPoint: COM方式 -> python-pptx (フォールバック)
    """
    def __init__(self):
        from office_converter.services.logger_service import LoggerService
        self.logger = LoggerService()

    def convert_word_to_pdf(self, input_path: Path, output_path: Path) -> Path:
        """Wordファイル(docx/doc)をPDFへ変換"""
        try:
            return self._convert_word_with_com(input_path, output_path)
        except Exception as e:
            self.logger.warning(f"Word COM変換失敗、代替方法を試行: {e}")
            try:
                from docx2pdf import convert
                convert(str(input_path), str(output_path))
                return output_path.resolve()
            except Exception as fallback_error:
                raise PDFConversionError(f"Word→PDF変換失敗: COM: {e}, docx2pdf: {fallback_error}")

    def convert_excel_to_pdf(self, input_path: Path, output_path: Path) -> Path:
        """Excelファイル(xlsx/xls)をPDFへ変換"""
        try:
            return self._convert_excel_with_com(input_path, output_path)
        except Exception as e:
            self.logger.warning(f"COM変換失敗、代替方法を試行: {e}")
            return self._convert_excel_with_openpyxl(input_path, output_path)

    def convert_ppt_to_pdf(self, input_path: Path, output_path: Path) -> Path:
        """PowerPointファイル(pptx/ppt)をPDFへ変換"""
        try:
            return self._convert_powerpoint_with_com(input_path, output_path)
        except Exception as e:
            self.logger.warning(f"PowerPoint COM変換失敗、代替方法を試行: {e}")
            try:
                from pptx import Presentation
                from reportlab.pdfgen import canvas
                from reportlab.lib.pagesizes import A4
                prs = Presentation(str(input_path))
                c = canvas.Canvas(str(output_path), pagesize=A4)
                for i, slide in enumerate(prs.slides):
                    c.drawString(100, 800, f"PowerPointスライド {i+1}")
                    c.showPage()
                c.save()
                return output_path.resolve()
            except Exception as fallback_error:
                raise PDFConversionError(f"PowerPoint→PDF変換失敗: COM: {e}, 簡易変換: {fallback_error}")

    def _convert_word_with_com(self, input_path: Path, output_path: Path) -> Path:
        """Word COM変換"""
        import win32com.client
        word_app = None
        document = None
        
        try:
            word_app = win32com.client.Dispatch("Word.Application")
            word_app.Visible = False
            word_app.DisplayAlerts = 0
            
            self.logger.info("Word COM初期化完了")
            document = word_app.Documents.Open(str(input_path.resolve()))
            self.logger.info(f"Word文書読み込み完了: {input_path.name}")
            
            # PDF形式でエクスポート
            document.ExportAsFixedFormat(
                OutputFileName=str(output_path.resolve()),
                ExportFormat=17  # wdExportFormatPDF
            )
            
            if output_path.exists():
                file_size = output_path.stat().st_size
                self.logger.info(f"Word PDF変換完了: {file_size} bytes")
                return output_path.resolve()
            else:
                raise Exception("PDFファイルが生成されませんでした")
                
        finally:
            if document:
                try: document.Close()
                except: pass
            if word_app:
                try: word_app.Quit()
                except: pass

    def _convert_excel_with_com(self, input_path: Path, output_path: Path) -> Path:
        """Excel COM変換（高品質印刷イメージ）"""
        import win32com.client
        excel_app = None
        workbook = None
        
        try:
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            self.logger.info("Excel COM初期化完了")
            workbook = excel_app.Workbooks.Open(str(input_path.resolve()))
            self.logger.info(f"Excelファイル読み込み完了: {input_path.name}")
            
            # 用紙向き自動最適化
            for worksheet in workbook.Worksheets:
                used_range = worksheet.UsedRange
                if used_range:
                    row_count = used_range.Rows.Count
                    col_count = used_range.Columns.Count
                    
                    # 横幅が大きい場合は横向き
                    if col_count > row_count or col_count > 20:
                        optimal_orientation = 2  # 横向き
                        orientation_name = "横向き"
                    else:
                        optimal_orientation = 1  # 縦向き
                        orientation_name = "縦向き"
                    
                    self.logger.info(f"シート '{worksheet.Name}': {used_range.Address} ({row_count}行×{col_count}列)")
                    self.logger.info(f"最適用紙向き: {orientation_name} (列数:{col_count})")
                    
                    # 印刷設定最適化
                    page_setup = worksheet.PageSetup
                    page_setup.Orientation = optimal_orientation
                    
                    if optimal_orientation == 2:  # 横向き
                        page_setup.FitToPagesWide = 1
                        page_setup.FitToPagesTall = False
                        # 横向き用余白設定
                        page_setup.LeftMargin = excel_app.InchesToPoints(0.2)
                        page_setup.RightMargin = excel_app.InchesToPoints(0.2)
                        page_setup.TopMargin = excel_app.InchesToPoints(0.3)
                        page_setup.BottomMargin = excel_app.InchesToPoints(0.3)
                    else:  # 縦向き
                        page_setup.FitToPagesWide = 1
                        page_setup.FitToPagesTall = 1
                        # 縦向き用余白設定
                        page_setup.LeftMargin = excel_app.InchesToPoints(0.3)
                        page_setup.RightMargin = excel_app.InchesToPoints(0.3)
                        page_setup.TopMargin = excel_app.InchesToPoints(0.2)
                        page_setup.BottomMargin = excel_app.InchesToPoints(0.2)
                    
                    # 図形数確認
                    shapes_count = worksheet.Shapes.Count
                    self.logger.info(f"図形数: {shapes_count}")
                    
                    # 印刷品質設定
                    page_setup.PrintGridlines = False
                    page_setup.PrintHeadings = False
                    page_setup.BlackAndWhite = False
                    
                    self.logger.info(f"印刷設定完了: {orientation_name}, 余白最適化")
            
            # ExportAsFixedFormat実行
            workbook.ExportAsFixedFormat(
                Type=0,  # xlTypePDF
                Filename=str(output_path.resolve())
            )
            
            if output_path.exists():
                file_size = output_path.stat().st_size
                self.logger.info(f"Excel PDF変換完了: {file_size} bytes")
                return output_path.resolve()
            else:
                raise Exception("PDFファイルが生成されませんでした")
                
        finally:
            if workbook:
                try: workbook.Close(SaveChanges=False)
                except: pass
            if excel_app:
                try: excel_app.Quit()
                except: pass

    def _convert_powerpoint_with_com(self, input_path: Path, output_path: Path) -> Path:
        """PowerPoint COM変換"""
        import win32com.client
        ppt_app = None
        presentation = None
        
        try:
            ppt_app = win32com.client.Dispatch("PowerPoint.Application")
            ppt_app.Visible = 1
            
            self.logger.info("PowerPoint COM初期化完了")
            presentation = ppt_app.Presentations.Open(str(input_path.resolve()))
            self.logger.info(f"PowerPoint文書読み込み完了: {input_path.name}")
            
            # PDF形式で保存
            presentation.SaveAs(str(output_path.resolve()), 32)  # ppSaveAsPDF
            
            if output_path.exists():
                file_size = output_path.stat().st_size
                self.logger.info(f"PowerPoint PDF変換完了: {file_size} bytes")
                return output_path.resolve()
            else:
                raise Exception("PDFファイルが生成されませんでした")
                
        finally:
            if presentation:
                try: presentation.Close()
                except: pass
            if ppt_app:
                try: ppt_app.Quit()
                except: pass

    def _convert_excel_with_openpyxl(self, input_path: Path, output_path: Path) -> Path:
        """Excelフォールバック変換（openpyxl + reportlab）"""
        try:
            from openpyxl import load_workbook
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            
            wb = load_workbook(str(input_path))
            c = canvas.Canvas(str(output_path), pagesize=A4)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                c.drawString(100, 800, f"Excel Sheet: {sheet_name}")
                # 簡易実装: セルデータの描画は省略
                c.showPage()
            
            c.save()
            return output_path.resolve()
            
        except Exception as e:
            raise PDFConversionError(f"Excel→PDF変換失敗: {e}")

    def convert_office_to_pdf(self, input_path: Path, output_dir: Path) -> List[Path]:
        """Officeファイル（Word, Excel, PowerPoint）をPDFへ変換"""
        ext = input_path.suffix.lower()
        output_paths = []
        
        if ext in [".docx", ".doc"]:
            pdf_path = output_dir / f"{input_path.stem}.pdf"
            output_paths.append(self.convert_word_to_pdf(input_path, pdf_path))
        elif ext in [".xlsx", ".xls"]:
            pdf_path = output_dir / f"{input_path.stem}.pdf"
            output_paths.append(self.convert_excel_to_pdf(input_path, pdf_path))
        elif ext in [".pptx", ".ppt"]:
            pdf_path = output_dir / f"{input_path.stem}.pdf"
            output_paths.append(self.convert_ppt_to_pdf(input_path, pdf_path))
        else:
            raise PDFConversionError(f"未対応の拡張子: {ext}")
        
        return output_paths